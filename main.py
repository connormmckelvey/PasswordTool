from tkinter import *
from tkinter import messagebox
import random
import string
import pyperclip as pc
from openpyxl import Workbook, load_workbook

#varribles //////////////
password = ""
alphabet = "aqzxswedcvfrtgbnhyujmkilop"
numbers = "1234567890"
syms = "!@#$%^&*()_+{}:?><|-=][;'/.,']"
characters = alphabet + numbers + syms
alpha_numbers = alphabet + numbers
gen_psw = False
usesyms = False
add_xlsx = False
root = ''
psw_length = 8
application = 'google'
button1_name = 'Generate & Save'


cell_txt_file = open(r"line_log.txt","r")
#returns lines as list 
cell_numbs = cell_txt_file.readlines()
# [0] app_cell_numb
# [1] password_cell_numb
app_cell_numb = int(cell_numbs[0])
password_cell_numb = int(cell_numbs[1])
cell_txt_file.close()
print(app_cell_numb,password_cell_numb)

#FIRST TIME OPENED
if app_cell_numb and password_cell_numb == 1:
    #welcome window declarations
    welc = Tk()
    welc.geometry("500x220")
    welc.title("Welcome to Password Tool!")

    #welc gui varribles
    welc_title_lbl = Label(welc,fg='purple',text='Welcome!',font= ('Fixedsys', 40, 'bold'))
    welc_txt1_lbl = Label(welc,text='Password Tool Developed By Krazy Studios',font= ('arial',12, 'bold'))
    welc_txt2_lbl = Label(welc,fg='purple',text='Setup: Before using the program create a file called\n \"passwords.xlsx" in the same folder as this file.',font= ('arial', 12))
    welc_txt3_lbl = Label(welc,fg='purple',text='--------------------------\nEnjoy simply using, generating and saving passwords!\n--------------------------',font= ('arial', 12, 'bold'))
    welc_txt4_lbl = Label(welc,text='Password Tool is created by Krazy Studios, for more info check out README.txt',font= ('arial', 8)).place(x=10,y=190)

    #packing gui for welc
    welc_title_lbl.pack()
    welc_txt1_lbl.pack()
    welc_txt2_lbl.pack()
    welc_txt3_lbl.pack()

if app_cell_numb and password_cell_numb == 100:
    messagebox.showinfo('Wow', 'Wow you hit 100 entrys, thats way to many...')

def xlsx():
    global add_xlsx
    if add_xlsx == False:
        print('add_xlsx = true')
        add_xlsx = True
    else:
        add_xlsx = False

def syms_chb():
    global usesyms
    if usesyms == False:
        usesyms = True
    else:
        usesyms = False 

def genpsw():
    global gen_psw
    if gen_psw == False:
        gen_psw = True
        button1_name = 'Generate and Save'
    else:
        gen_psw = False

def gen():
    print("Function Run")
    global password
    password = ''
    application = app_ent.get()
    global app_cell_numb
    global password_cell_numb
    global add_xlsx
    global app_cell
    global password_cell
#generate password
    if gen_psw == True:
        for i in range(psw_length):
            print("for loop run")
            if usesyms == True:
                digit = random.choice(characters)
                password = password + digit
                print("password assigned IF")
            if usesyms == False:
                digit = random.choice(alpha_numbers)
                password = password + digit
                print("password:" + str(password))
    if gen_psw == False:
        password = password_ent.get()
    
    #add to xlsx
    
    if add_xlsx == True:
        
        app_cell = f'A{str(app_cell_numb)}'
        password_cell = f'B{str(password_cell_numb)}'
        print('added to xlsx file')
        #actions
        ws[app_cell].value = application
        ws[password_cell].value = password
        
        password_cell_numb += 1
        app_cell_numb += 1
        

        print(ws[password_cell].value)
        print(ws[app_cell].value)
        
        #SAVING ALL ITEMS TO FILES
        #xlsx password saves
        wb.save("passwords.xlsx")

        #txt cell numbs save
        cell_txt_file = open(r"line_log.txt","w")
        cell_numbs = [str(f"{app_cell_numb}\n"),str(password_cell_numb)]
        cell_txt_file.writelines(cell_numbs)
        cell_txt_file.close()

    label.config(text=f'Saved! {password} for {application}')

def copy():
    print(f"copying {password}")
    if password == "":
        messagebox.showinfo('No Password Copyed', 'You have not saved a password yet, so we couldn\'t copy your password')
    pc.copy(password)

def reset():
    res = messagebox.askokcancel('Reset All Data', 'Are You sure you would like to clear all data?')
    if res == True:
        print("Clearing all Data")
        #clear xlsx files
        for row in range(app_cell_numb):
            ws[f"A{row + 1}"].value = ""
            ws[f"B{row + 1}"].value = ""
            print(f"cleared A{row + 1} and B{row + 1}")
        wb.save("passwords.xlsx")


#xlsx declare
wb = load_workbook('passwords.xlsx')
ws = wb.active

#main window declarations
root = Tk()
root.geometry("500x350")
root.title("Password Tool")


#declare gui varribles ///////////////
title_lbl = Label(root,fg='purple',text='Password Tool',font= ('Fixedsys', 40, 'bold'))
app_lbl = Label(root,text='Application or Website:')
app_ent = Entry(root,)
password_lbl = Label(root,text='Password (leave blank for random)')
password_ent = Entry(root,)
add_xlsx_chb = Checkbutton(root,text='Add to Exell File?',command=xlsx)
gen_psw_chb = Checkbutton(root,text='Generate Random Password',command=genpsw)
usesyms_chb = Checkbutton(root,text='Include Syms',command=syms_chb)
label = Label(root, font = ('arial', 12, 'bold'))
button1 = Button(root,text=button1_name, font = ('arial', 12, 'bold'),command= gen)
button2 = Button(root,text="Copy to Clipboard", font =('arial',12,'bold'),command=copy)
button3 = Button(root,text="Reset All", font=('arial',10),command=reset).place(x=10,y=300)
file_location_lbl = Label(root,text='Before Running: create file \'passwords.xlsx\' in the same folder as this program.').place(x=10,y=330)

#pack gui
title_lbl.pack()
app_lbl.pack()
app_ent.pack()
password_lbl.pack()
password_ent.pack()
add_xlsx_chb.pack()
gen_psw_chb.pack()
usesyms_chb.pack()
label.pack()
button1.pack()
button2.pack()

root.mainloop()
